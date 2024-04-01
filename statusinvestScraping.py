import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import json

'''
FUNCIONALIDADES:

--le o arquivo de parametros (parametros.json) contendo os nomes das acoes para realizar a busca das acoes no site (https://statusinvest.com.br)

--realiza a raspagem dos dados selecionados no codigo de cada acao e os printa no console, tambem criando um arquivo .xlsx com os mesmos dados
 
--Filtra as colunas criando e mantendo apenas as necessarias para uma consulta mais clara dos dados das acoes
'''

# Declaracao dos arrays para armazenar as informacoes coletadas
valorAcao = []
divYield = []
vDividendo = []
pl = []
divLiquida = []
mrgLiquida = []
rOE = []
pvp = []
vpa = []
lpa = []
dyhisttorico = []
sAtuacao = []
nomeAcao = []
cagr5a = []
valorMercado = []
dPA = []
valuationBazin = []
descontoBazin = []
valuationGraham = []
descontoGraham = []
valuationGordon = []
descontoGordon = []
peg = []
payout = []
crescimentoEsperado = []
mediaCrescimento = []

class StatusInvestScraper:
    def __init__(self):
        self.driver = webdriver.Chrome()


    def tratStr(self, string_, pos):
        string = string_.split()
        return string[pos]

    def adcColor(self, dado, param):
        if dado < param:
            return '\033[91m'  # Vermelho
        else:
            return '\033[92m'  # Verde

    def obter_dados_acao(self, acao):
        global valorAcao,divYield,vDividendo,pl,divLiquida,mrgLiquida,rOE,pvp,vpa,lpa,dyhisttorico,sAtuacao,nomeAcao,cagr5a,\
            valorMercado,dPA,valuationBazin,descontoBazin,valuationGraham,descontoGraham,valuationGordon,descontoGordon,\
            peg,payout,crescimentoEsperado,mediaCrescimento

        driver = self.driver
        url = f"https://statusinvest.com.br/acoes/{acao}"
        driver.get(url)
        dy_color = '\033[91m'

        # Busca os dados do Dividend Yield
        try:
            dividend_yield_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="main-2"]/div[2]/div/div[1]/div/div[4]'))
            )
            dividend_yield = dividend_yield_element.text
            vl_dividend = float(self.tratStr(dividend_yield, 9).replace(",", "."))
            dividend_yield = float(self.tratStr(dividend_yield, 3).replace(",", "."))
            dy_color = self.adcColor(round(dividend_yield), 6)
            divYield.append(dividend_yield)
        except Exception as e:
            dividend_yield = 0
            divYield.append(dividend_yield)

        # Busca os dados do P/L
        p_l = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[1]/div/div[2]/div/div/strong'))
        )
        p_l = float(p_l.text.replace(",", "."))
        pl_color = self.adcColor(10, round(abs(p_l)))
        pl.append(p_l)
        if p_l < 0:
            pl_color = '\033[91m'

        # Busca os dados da Dívida Líquida
        divida_liquida = driver.find_element(By.XPATH,
                                             '//*[@id="indicators-section"]/div[2]/div/div[2]/div/div[2]/div').text
        try:
            divida_liquida = self.tratStr(divida_liquida, 3).replace(",", ".")
            dl_color = self.adcColor(3.5, round(float(divida_liquida)))
            divLiquida.append(divida_liquida)
            if float(divida_liquida) < 0:
                dl_color = '\033[91m'
        except:
            divida_liquida = "Indefinido"
            dl_color = '\033[0m'
            divLiquida.append(divida_liquida)

        # Busca os dados da Margem Líquida
        margem_liquida = driver.find_element(By.XPATH,
                                             '//*[@id="indicators-section"]/div[2]/div/div[3]/div/div[4]/div').text
        margem_liquida = self.tratStr(margem_liquida, 3).replace(",", ".").replace("%", "")
        ml_color = self.adcColor(round(float(margem_liquida)), 10)
        mrgLiquida.append(margem_liquida)

        # Busca os dados do ROE
        roe = driver.find_element(By.XPATH, '//*[@id="indicators-section"]/div[2]/div/div[4]/div/div[1]/div').text
        roe = self.tratStr(roe, 2).replace(",", ".").replace("%", "")
        floatRoe = float(roe)
        rOE.append(floatRoe)
        roe_color = self.adcColor(float(roe), 15)

        # Busca os dados do P/VP
        p_vp = driver.find_element(By.XPATH,
                                   '//*[@id="indicators-section"]/div[2]/div/div[1]/div/div[4]/div/div/strong').text
        pvp.append(p_vp)
        pvp_color = self.adcColor(1.5, (float(p_vp.replace(",", "."))))

        # Busca os dados do S. Atuacao
        s_atuacao = driver.find_element(By.XPATH,
                                        '//*[@id="company-section"]/div[1]/div/div[3]/div/div[1]/div/div/div/a/strong').text
        sAtuacao.append(s_atuacao)

        # Busca os dados do VPA
        vpa_element = driver.find_element(By.XPATH,
                                        '//*[@id="indicators-section"]/div[2]/div/div[1]/div/div[9]/div/div/strong').text
        vpaCorrigido = vpa_element.replace(',', '.')
        vpa_float = float(vpaCorrigido)
        vpa.append(vpa_float)

        # Busca os dados do LPA
        lpa_element = driver.find_element(By.XPATH,
                                        '//*[@id="indicators-section"]/div[2]/div/div[1]/div/div[11]/div/div/strong').text
        lpaCorrigido = lpa_element.replace(',', '.')
        lpa_float = float(lpaCorrigido)
        lpa.append(lpa_float)

        # Busca os dados do Valor ação
        valor_acao = driver.find_element(By.XPATH,
                                          '//*[@id="main-2"]/div[2]/div/div[1]/div/div[1]/div/div[1]/strong').text
        acaoCorrigido = valor_acao.replace(',', '.')
        valorAcao_float = float(acaoCorrigido)
        valorAcao.append(valorAcao_float)

        # Busca os dados do Nome ação
        nome_acao = driver.find_element(By.XPATH,
                                          '//*[@id="main-header"]/div[2]/div/div[1]/h1/small').text
        nomeAcao.append(nome_acao)

        # Busca os dados do CAGR LUCROS 5 ANOS
        cagr = driver.find_element(By.XPATH,
                                          '//*[@id="indicators-section"]/div[2]/div/div[5]/div/div[2]/div/div/strong').text
        replaceCagr = cagr.replace('%', '').replace(',', '.').replace("-", "0")
        float_cagr = float(replaceCagr)
        cagr5a.append(float_cagr)

        # Busca os dados do valor de mercado
        valor_mercado = driver.find_element(By.XPATH,
                                          '//*[@id="company-section"]/div[1]/div/div[2]/div[7]/div/div/strong').text
        valor_mercado = valor_mercado.replace('.', '')
        float_valorMercado = float(valor_mercado)
        valorMercado.append(float_valorMercado)


        #  faz o scroll da pag ate embaixo
        # ======================================
        SCROLL_PAUSE_TIME = 0.5
        # Get scroll height
        last_height = driver.execute_script("return document.body.scrollHeight")

        while True:
            # Scroll down to bottom
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

            # Wait to load page
            time.sleep(SCROLL_PAUSE_TIME)

            # Calculate new scroll height and compare with last scroll height
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        time.sleep(10)
        #  FIM faz o scroll da pag ate embaixo

        # Busca os dados do Payout
        find_payout = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//*[@id='payout-section']/div/div/div[1]/div[1]/div[2]/strong/span")))
        valu_payout = find_payout.get_attribute("innerHTML")
        valu_payout = valu_payout.replace('%', '').replace(',', '.')
        valu_payout = float(valu_payout)
        payout.append(valu_payout)

        # Busca os dados do D.Y -historico
        driver.get(url)
        time.sleep(3)
        driver.find_element(By.XPATH, '//*[@id="indicators-section"]/div[1]/div[2]/button[2]').click()
        time.sleep(3)
        table = driver.find_element(By.XPATH, '//*[@id="indicators-section"]/div[3]/div[2]/div/div[2]/div')
        soup = BeautifulSoup(table.get_attribute('outerHTML'), "html.parser")

        # Trata os dados do D.Y -historico
        table_headers = []
        for th in soup.find_all('div', 'th'):
            table_headers.append(th.text)
            if th.text < "2019":
                break

        # Trata os dados do D.Y -historico
        table_data = []
        for row in soup.find_all('div', 'tr'):
            columns = row.find_all('div', 'td')
            output_row = []
            count = 0
            for column in columns:
                output_row.append(column.text)
                if count > len(table_headers) - 2:
                    break
                count = count + 1
            table_data.append(output_row)

        # Trata os dados do D.Y -historico
        for i in range(len(table_data[1])):
            # Remove o caractere '%' e substitui ',' por '.'
            numero_sem_percentual = table_data[1][i].replace('%', '').replace(',', '.').replace("-", "0")

            # Converte a string resultante para float
            numero_float = float(numero_sem_percentual)

            # Atualiza o elemento na lista
            table_data[1][i] = numero_float

        # Trata os dados do D.Y -historico
        soma = 0
        num_anos = len(table_data[1])
        for data in table_data[1]:
            soma += data
        media_dy = (soma / num_anos).__round__(2)
        dyhisttorico.append(media_dy)

        titulos = {
            'Valor ação': valor_acao,
            'dy_color': dy_color,
            'Dividend Yield': dividend_yield,
            'V. Dividendo': vl_dividend,
            'pl_color': pl_color,
            'P/L': p_l,
            'dl_color': dl_color,
            'Dívida Líquida': divida_liquida,
            'ml_color': ml_color,
            'Margem Líquida': margem_liquida,
            'roe_color': roe_color,
            'ROE': roe,
            'pvp_color': pvp_color,
            'P/VP': p_vp,
            'color_reset': '\033[0m',
            'VPA': vpa_element,
            'LPA': lpa_element,
            'D.Y -historico': media_dy,
            'S. Atuacao': s_atuacao
        }

        #------
        # Trata os dados obtidos:
        #------
        #RISCO em %
        risco = 0.15

        # DPA usando a formula:
        try:
            dpa = (valorAcao_float * (dividend_yield / 100)).__round__(2)
            dPA.append(dpa)
        except:
            dpa = 0
            dPA.append(dpa)

        # Valuation Bazin usando a formula:
        try:
            valu_bazin = (dpa / 0.06).__round__(2)
            valuationBazin.append(valu_bazin)
        except:
            valu_bazin = 0
            valuationBazin.append(valu_bazin)

        # Desconto Bazin usando a formula:
        try:
            desc_bazin = (((valu_bazin - valorAcao_float) / valu_bazin) * 100).__round__(2)
            descontoBazin.append(desc_bazin)
        except:
            desc_bazin = 0
            descontoBazin.append(desc_bazin)

        # Valuation Graham usando a formula:
        try:
            valu_graham = ((22.5 * lpa_float * vpa_float) ** 0.5).__round__(2)
            valuationGraham.append(valu_graham)
        except:
            valu_graham = 0
            valuationGraham.append(valu_graham)

        # Desconto Graham usando a formula:
        try:
            desc_graham = (((valu_graham - valorAcao_float) / valu_graham) * 100).__round__(2)
            descontoGraham.append(desc_graham)
        except:
            desc_graham = 0
            descontoGraham.append(desc_graham)

        # Valuation Gordon usando a formula:
        try:
            valu_gordon = ((dpa * (1 + (float_cagr / 100))) / risco).__round__(2)
            valuationGordon.append(valu_gordon)
        except:
            valu_gordon = 0
            valuationGordon.append(valu_gordon)

        # Desconto Gordon usando a formula:
        try:
            desc_gordon = (((valu_gordon - valorAcao_float) / valu_gordon) * 100).__round__(2)
            descontoGordon.append(desc_gordon)
        except:
            desc_gordon = 0
            descontoGordon.append(desc_gordon)

        # PEG usando a formula:
        try:
            valu_peg = (p_l / (float_cagr)).__round__(2)
            peg.append(valu_peg)
        except:
            valu_peg = 0
            peg.append(valu_peg)

        # Crescimento Esperado usando a formula:
        try:
            valu_cresEsp = (((1 - (valu_payout / 100)) * (floatRoe / 100)) * 100).__round__(2)
            crescimentoEsperado.append(valu_cresEsp)
        except:
            valu_cresEsp = 0
            crescimentoEsperado.append(valu_cresEsp)

        # Media de Crescimento usando a formula:
        try:
            valu_med = ((valu_cresEsp + float_cagr) / 2).__round__(2)
            mediaCrescimento.append(valu_med)
        except:
            valu_med = 0
            mediaCrescimento.append(valu_med)


        return titulos

def montaExcel():
    global valorAcao, divYield, vDividendo, pl, divLiquida, mrgLiquida, rOE, pvp, vpa, lpa, dyhisttorico, sAtuacao, nomeAcao, cagr5a, valorMercado, dPA,\
        valuationBazin,descontoBazin,valuationGraham,descontoGraham,valuationGordon,descontoGordon,peg,payout,crescimentoEsperado,mediaCrescimento

    # Prepara a planilha na memoria
    wb_Saida = openpyxl.Workbook()
    sheet_Saida = wb_Saida.create_sheet('CDB')
    wb_Saida.active = sheet_Saida

    acoesPage = wb_Saida['CDB']

    # Cria 5 linhas em brando no inicio
    acoesPage.cell(row=5, column=14, value="")

    # Define os valores da primeira linha
    acoesPage.append(["acoes: ", "Nome da ação: ", "S. Atuacao: ", "Valor ação: ", "Dívida Líquida: ", "Margem Líquida: ", "Valuation Bazin: ", "Desconto Bazin: ", "Valuation Graham: ",
                      "Desconto graham: ", "Valuation Gordon: ", "Desconto Gordon: ", "PEG: ", "Dividend Yield: ", "P/L: ", "LPA: ", "VPA: ", "DPA: ", "P/VP: ",
                      "Payout: ", "ROE: ", "CAGR lucros 5 anos: ", "D.Y -historico: ", "Crescimento Esperado", "Media de Crescimento: ", "Valor de Mercado: "])

    # Define os valores das demais linhas atraves do loop
    for i in range(len(acoes)):
        acoesPage.append(
            [acoes[i], nomeAcao[i], sAtuacao[i], valorAcao[i], divLiquida[i], mrgLiquida[i], valuationBazin[i], descontoBazin[i], valuationGraham[i],
             descontoGraham[i], valuationGordon[i], descontoGordon[i], peg[i], divYield[i], pl[i], lpa[i], vpa[i], dPA[i], pvp[i],
             payout[i], rOE[i], cagr5a[i], dyhisttorico[i], crescimentoEsperado[i], mediaCrescimento[i], valorMercado[i]])

    # Grava a saida .xlsx
    wb_Saida.save('statusinvest.xlsx')



if __name__ == '__main__':
    with open("parametros.json", 'r') as arquivo:
        dados = json.load(arquivo)

    #Adiciona os parametros em um array
    acoesStr = dados.get('acoes', '')
    acoes = acoesStr.split(', ')


    scraper = StatusInvestScraper()


    # Printa no console os valores obtidos
    for acao in acoes:

        dados_acao = scraper.obter_dados_acao(acao)

        if dados_acao is not None:
            print(f"\n\nDados da ação {acao}:")
            chave = []
            valor = []

            for chave, valor in dados_acao.items():
                if "color" not in chave:
                    print(f" |{chave.ljust(14)} |", end="")
                else:
                    print(f"{valor}", end="")

            print()
            for chave, valor in dados_acao.items():
                if "color" not in chave:
                    print(f" |{str(valor).ljust(14)} |", end="")
                else:
                    print(f"{valor}", end="")

#Executa o excel
    montaExcel()
